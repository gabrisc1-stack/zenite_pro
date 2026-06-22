from playwright.sync_api import sync_playwright
from openpyxl import load_workbook
import re

LOGIN_JSON = "login.json"

ARQUIVO_EXCEL = r"app_zenite\PRECIFICAÇÃO IPHONE.xlsx"

produtos = [
    {"modelo": "iPhone 15", "memoria": "128GB"},
    {"modelo": "iPhone 15 Plus", "memoria": "128GB"},
    {"modelo": "iPhone 15 Pro", "memoria": "128GB"},
    {"modelo": "iPhone 15 Pro Max", "memoria": "256GB"},

    {"modelo": "iPhone 16", "memoria": "128GB"},
    {"modelo": "iPhone 16 Plus", "memoria": "128GB"},
    {"modelo": "iPhone 16 Pro", "memoria": "128GB"},
    {"modelo": "iPhone 16 Pro", "memoria": "256GB"},
    {"modelo": "iPhone 16 Pro Max", "memoria": "256GB"},
    {"modelo": "iPhone 16 Pro Max", "memoria": "512GB"},

    {"modelo": "iPhone 17", "memoria": "256GB"},
    {"modelo": "iPhone 17 Pro", "memoria": "256GB"},
    {"modelo": "iPhone 17 Pro", "memoria": "512GB"},
    {"modelo": "iPhone 17 Pro", "memoria": "1TB"},
    {"modelo": "iPhone 17 Pro Max", "memoria": "256GB"},
    {"modelo": "iPhone 17 Pro Max", "memoria": "512GB"},
    {"modelo": "iPhone 17 Pro Max", "memoria": "1TB"},
    {"modelo": "iPhone 17 Pro Max", "memoria": "2TB"},

    {"modelo": "Apple Watch SE 2 44mm", "memoria": ""},
    {"modelo": "Apple Watch Se 3 40", "memoria": ""},
    {"modelo": "Apple Watch Se 3 44", "memoria": ""},
    {"modelo": "Apple Watch S9 41mm", "memoria": ""},
    {"modelo": "Apple Watch S9 45mm", "memoria": ""},
    {"modelo": "Apple Watch S11 42mm", "memoria": ""},
    {"modelo": "Apple Watch S11 46mm", "memoria": ""},
    {"modelo": "Apple Watch Ultra 49mm", "memoria": ""},
    {"modelo": "Apple Watch Ultra 2 49mm", "memoria": ""},

    {"modelo": "AirPods 4 ANC", "memoria": ""},
    {"modelo": "AirPods 4", "memoria": ""},
    {"modelo": "AirPods Pro 1", "memoria": ""},
    {"modelo": "AirPods Pro 2", "memoria": ""},
    {"modelo": "AirPods Pro 3", "memoria": ""},

    {"modelo": "MacBook Air M1", "memoria": "256GB"},
    {"modelo": "MacBook Air M1", "memoria": "512GB"},
    {"modelo": "MacBook Neo 13", "memoria": "256GB"},
    {"modelo": "MacBook Neo 13", "memoria": "512GB"},
    {"modelo": "MacBook Air M4", "memoria": "256GB"},
    {"modelo": "MacBook Air M4", "memoria": "512GB"},
    {"modelo": "MacBook Air M4", "memoria": "1TB"},
    {"modelo": "MacBook Air M5", "memoria": "256GB"},
    {"modelo": "MacBook Air M5", "memoria": "512GB"},
    {"modelo": "MacBook Air M5", "memoria": "1TB"},
    {"modelo": "MacBook Pro M1", "memoria": "256GB"},
    {"modelo": "MacBook Pro M1", "memoria": "512GB"},
    {"modelo": "MacBook Pro M1", "memoria": "1TB"},
    {"modelo": "MacBook Pro M4", "memoria": "256GB"},
    {"modelo": "MacBook Pro M4", "memoria": "512GB"},
    {"modelo": "MacBook Pro M4", "memoria": "1TB"},
]

def extrair_precos(texto):
    encontrados = re.findall(r"\d{1,2}\.\d{3}", texto)
    valores = []

    for preco in encontrados:
        try:
            valor = int(preco.replace(".", ""))
            if valor > 500:
                valores.append(valor)
        except:
            pass

    return valores

resultados = []

with sync_playwright() as p:
    browser = p.chromium.launch(headless=True)
    context = browser.new_context(storage_state=LOGIN_JSON)
    page = context.new_page()

    page.goto("https://appdamaca.com.br/fornecedores", timeout=60000)
    page.wait_for_timeout(4000)

    for produto in produtos:
        modelo = produto["modelo"]
        memoria = produto["memoria"]
        pesquisa = modelo
        print(f"Buscando: {pesquisa}")

        try:
            busca = page.locator('input[placeholder*="Buscar"]').first
            busca.fill("")
            page.wait_for_timeout(500)

            busca.fill(pesquisa)
            page.wait_for_timeout(3000)

            try:
                if "iphone" in modelo.lower() or "macbook" in modelo.lower():
                    page.locator("select").nth(1).select_option(label=memoria)
                else:
                    page.locator("select").nth(1).select_option(label="Todos")

                page.wait_for_timeout(2000)
            except:
                pass

            cards = page.locator("body").inner_text().split("\n")

            texto_filtrado = ""

            for i, linha in enumerate(cards):
                bloco = "\n".join(cards[i:i+12]).lower()

                if "usado" in bloco or "cpo" in bloco or "seminovo" in bloco:
                    continue

                if modelo.lower() == "iphone 17":
                    bloco_limpo = bloco.lower()
                    bloco_limpo = bloco_limpo.replace(" ", "")
                    bloco_limpo = bloco_limpo.replace("-", "")

                    if "17e" in bloco_limpo:
                        continue

                    if "iphone17pro" in bloco_limpo:
                        continue
                    
                    if "iphone17promax" in bloco_limpo:
                        continue
                    
                    if "iphone17" not in bloco_limpo:
                        continue

                    if memoria.lower() not in bloco.lower():
                        continue

                texto_filtrado += "\n" + "\n".join(cards[i:i+12])

            texto = texto_filtrado
            valores = extrair_precos(texto)

            if valores:
                menor_preco = min(valores)
            else:
                menor_preco = 0

        except Exception as e:
            menor_preco = "Erro"
            print("ERRO REAL:", e)

        resultados.append({
            "Modelo": modelo,
            "GB/TB": memoria,
            "Menor Preço": menor_preco
        })

        print(f"Menor preço encontrado: {menor_preco}")

print("TOTAL DE RESULTADOS:", len(resultados))

wb = load_workbook(ARQUIVO_EXCEL)
ws = wb.active

linha = 2

for item in resultados:
    ws[f"C{linha}"] = item["Menor Preço"]
    linha += 1

wb.save(ARQUIVO_EXCEL)

print("PLANILHA ATUALIZADA COM SUCESSO")